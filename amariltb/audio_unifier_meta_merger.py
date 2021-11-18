"""

INSTRUCTIONS:

1) in the RUN section - change variable 'target_dir' to be the result directory which the audio_unifier.py script had created:
(target_dir = 'result') 

2) when you have the sonix outputs (for the .wav files which you created in the audio_unifier.py script) 
put all sonix txt files into that directory but MAKE SURE that each file is named as follows:

if the meta file which was created for your 'wav' is called: 'free_telaviv_adhd_hitId___heb_1_meta.xlsx'
the corresponding sonix file must be named : 'free_telaviv_adhd_hitId___heb_1_sonix.txt' 
(just swap 'meta.xlsx' for 'sonix.txt')

3) in the 'target_dir' a file called  telaviv_adhd_pi.json will be created to be used later with parser.
(in order to produce the ProductWords output.)
for viewing the output you can set the variable 'create_xlsx_target' to True, 
and an xlsx file will be created as well.

"""

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import time
import os
import json
import ntpath

# TBD: move to separate const file which should be accesed globaly by all modules
Language_Code_Dict = {
    "heb":"hebrew",
    "en":"english"
}

Categories = {
    "animals":"animals",
    "food":"food",
    "plants":"plants", 
    "transport":"transport", 
    "drinks" : "drinks"
}

def xlsx_to_ar(filename):

    wb1 = load_workbook(filename)
    ar = []
    sheet =  wb1.worksheets[0]
    for row_cells in sheet.iter_rows():
        item = {}
        item['filename'] = row_cells[0].value
        item['intervalStart'] = float(row_cells[1].value)
        item['intervalEnd'] = float(row_cells[2].value)
        item['duration'] = float(row_cells[3].value)
        item['index'] = row_cells[4].value
        ar.append(item)

    return ar

def get_sec(time_str):
    
    """Get Seconds from time."""
    h, m, s = time_str.split(':')
    return int(h) * 3600 + int(m) * 60 + float(s)

def get_word(ar):
    ar.pop(0)
    word_str = ''

    for index,word in enumerate(ar):
        if(word == ''):
            continue

        if(index==0):
            word_str+= word
        else:
            word_str+=  '_' + word 

    return word_str.replace('.','').replace(',','')
        
# TBD : move to parser module
def parseSonixData(filename):
    parsed_data = []
    
    # get txt:
    #    file = open(filename,mode='r',encoding="utf8")

    file = open(filename,mode='r')
    txt_data = file.read()
    file.close()
    # parse:
    lines = txt_data.split('\n\n')
    for line_number,line in enumerate(lines):
        #skip first line:
        if '.wav' in line:
            continue

        # empty line ?
        if (line == ''):
            continue

        # has multiple words Error ?
        if( line.count(',') > 1 ):
            print('Transcript Warning: found several words in one item. \nitem:'+line  +' \n(file:'+filename+' - line number:'+str(line_number)+')') 
            is_error = input("stop the processing? (y/n)")
            if(is_error=='y'):
                raise AssertionError('Transcript Error: found several words in one item. item:'+line  +' (file:'+filename+' - line number:'+str(line_number)+')')  
        
        # parse line
        line_parts = line.split(' ')
        if(len(line_parts) < 2):
            print ('Error: line is not formatted as expected')
            print(line)
            continue

        # get ts in seconds:
        tsStr = line_parts[0].replace('[','').replace(']','')
        ts = get_sec(tsStr)
        
        
        word = get_word(line_parts)
       
        parsed_data.append({
            "text":word,
            "xmin": ts,
            "startStr": tsStr,
            "filename":path_leaf(filename)
        })

    return parsed_data

def path_leaf(path):
    head, tail = ntpath.split(path)
    return tail or ntpath.basename(head)

def update_sheet(filename,data):
    wb = Workbook()
    ws = wb.active 
    ws.title = 'result_meta'
    dest_filename = filename
    for data_item in data :
        ws.append([ data_item['text'],data_item['xmin'],data_item['filename'] ])
    
    wb.save(filename = dest_filename)
    return
def keyExists(dict, key):   
    if key in dict.keys(): 
        return True 
    else: 
        return False

# TBD: for filename get entire path
def createRecordingData(items,filename='',language = '',participant=''):
    return {
        'filename':filename,
        'assignmentId': os.path.splitext(filename)[0],
        'language':language,
        'items':items
    }


def create_parser_input_json(target_filename,data,experiment_info):
    # ask for category of this folder (experiment)
    # propmt user input :
    print( "please enter the relevant category for this experiment (folder):")
    user_category = str(input(""))
    # validate category:
    if(user_category not in  Categories.values() ):
        raise NameError('Category which was entered does not exist in categories.')

    json_data = {'recordings':[]}
    # data sorted by filename:
    recordings_dict = {}
    for data_item in data :

        # dont write the starts:
        if(data_item['word'] == 'start'):
            continue

        json_item = {
            'word':data_item['word'],
            'start':round(data_item['start'],3),
            'category':user_category
        }
        key = data_item['filename']
        # create array if does not exist:
        if(keyExists(recordings_dict, key)):
            ar = recordings_dict[key]
            ar.append(json_item)
            recordings_dict[key] = ar 
        else:
            ar = []
            ar.append(json_item)
            recordings_dict[key] = ar

    # dict to ar:
    for filename in recordings_dict.keys():
        items = recordings_dict[filename]
        recording_data = createRecordingData(items,filename=filename,language= experiment_info["language"])
        json_data['recordings'].append(recording_data)

    if("experiment_name" in experiment_info):
        json_data['experimentName'] = experiment_info["experiment_name"]

    with open(target_filename, 'w') as fp:
        json.dump(json_data, fp,sort_keys=True, indent=4)

def processSonixData(sonix_data,meta_data):
    result_data = []
    index = 0
    for sonix_item in sonix_data:
        is_new = False
        # find interval :
        interval = {}
 
        for meta_item in meta_data:

            # found interval ?
            if ( sonix_item['xmin'] > meta_item['intervalStart'] and 
                 sonix_item['xmin'] < meta_item['intervalEnd']   ):
                

                if(meta_item['index']!=index):
                    index = meta_item['index']
                    is_new = True
                
                interval = meta_item
                break
        
        # sanity:
        if(len(interval.values()) == 0):
            print ('Error: could not find interval for sonix_item:')
            print(sonix_item)
            continue
        else:
            # append start item if new user:
            if(is_new):
                start_item = {
                    'text':'start',
                    'xmin':'',
                    'startStr':'',
                    'filename':sonix_item['filename'],
                    'participant_id':interval['filename'],
                    'file_duration' : meta_item['intervalEnd'] - meta_item['intervalStart']
                }
                result_data.append(start_item) 
            
            sonix_item['index'] = interval['index']
            #sonix_item['filename'] = interval['filename']
            sonix_item['participant_id'] = interval['filename']
            sonix_item['xmin'] -= meta_item['intervalStart']
            sonix_item['file_duration'] = meta_item['intervalEnd'] - meta_item['intervalStart']

            #sonix_item['text'] = sonix_item['word']
            result_data.append(sonix_item)
    
    return result_data

def merge(meta_data_filename,sonix_data_filename):

    # get meta file:
    meta_data = xlsx_to_ar(meta_data_filename)

    # sort by index to make sure we are using the correct intervals:
    meta_data = sorted(meta_data, key=lambda x: x["index"], reverse=False)

    # get sonix data:
    sonix_data = parseSonixData(sonix_data_filename)

    # process sonix data: 
    amaril_data = processSonixData(sonix_data,meta_data)
    
    return amaril_data

def file_exists(filename,dir):
    for file in os.listdir(dir):
        if file == filename:
            return True
    return False

def merge_meta_into_sonix(source_dir,merged_filename):
    create_xlsx_target = True

    # iterate  sonix txt files and look for similar named meta files to merge with:
    total_ar = []

    for current_file in os.listdir(source_dir):
        if current_file.endswith(".txt"):
            #os.path.join(target_dir, file)
            sonix_data_filename = current_file
            meta_data_filename = sonix_data_filename.replace("sonix.txt","meta.xlsx")

            if(file_exists(meta_data_filename,source_dir)):
                meta_full_filename = os.path.join(source_dir, meta_data_filename)
                sonix_full_filename = os.path.join(source_dir, sonix_data_filename)
                data_ar = merge(meta_full_filename,sonix_full_filename)
                total_ar += data_ar
            
            else:
                print('Error: could not find meta file to couple with:',sonix_data_filename)

    # write to file:
    if(create_xlsx_target):
        update_sheet(merged_filename,total_ar)

    # create json:
    return total_ar
