import json
import codecs
import uuid
from decimal import Decimal
import hashlib
import json
from .pw_converter_input_normelizer import normelize_prrat,normelize_sonix_data
from .pw_converter_const import Categories,Languages,Tables
from .pw_converter_db import DynamoDB #get_index,get_transforms,update_index,update_transforms,update_product_words
from .pw_converter_utils import normelize_word,update_index_sheet
import os
from openpyxl import load_workbook
from openpyxl import Workbook

dynamodb = DynamoDB()

"""
# Note: for python 2.7 compatibility - uncomment the below:
try: 
    input = raw_input
except NameError: 
    pass
"""

def get_input(filename):
    # get file contents:
    #full_filename = os.path.join(directory, filename)
    return json.load(codecs.open(filename, 'r', 'utf-8-sig'), parse_float=Decimal)

def validate_input(input_data):
    # TBD: hard coded for now (ask user for missing details) experimentNane should appear in some Table ?
    # TBD: wordproducts only filename (or only assignmentId)?
    input_data['language'] = Languages["english"]
    input_data['experimentName'] = "experimentName" #TBD: no need
    input_data['category'] = Categories["animals"]
    return input_data    

def rev(s,langauge): 
    if(langauge == Languages["hebrew"]):
        return s[::-1]
    else:
        return s 

def get_index_dict_id(language,category):
    return 'index_'+language+'_'+category

def get_transforms_dict_id(language,category):
    return 'transforms_'+language+'_'+category

def get_index_dict(index_dicts,language,category):
    index_dict_id = get_index_dict_id(language,category)
    
    index_dict = None
    if(index_dict_id in index_dicts):
        index_dict = index_dicts[index_dict_id]
    else:
        index_dict = dynamodb.get_index(language,category)
        index_dicts[index_dict_id] = index_dict
    
    return index_dict

def index_dict_add_last_item(index_dicts,language,category,new_index):
    # update db:
    dynamodb.update_index([new_index])

    # updated cache :
    #index_dict = dynamodb.get_index(language,category)
    index_dict_id = get_index_dict_id(language,category)
    index_dict = index_dicts[index_dict_id]

    (word,index) = (new_index["word"] , new_index["id"] )
    index_dict[word] = index
    index_dict['nextIndexVal'] = index_dict['nextIndexVal'] + 1
    
    index_dicts[index_dict_id] = index_dict

def get_missing_languages(language):
    missing_languages = list(Languages.keys())
    missing_languages.remove(language)
    return missing_languages

def create_index_item(category,language,index,word_to_add):
    # create new index for transformed_word:
    id =  word_to_add + category + language
    
    hashed_id = hashlib.sha1(id.encode("UTF-8")).hexdigest()[5:]

    new_index = {
        "id" : hashed_id,
        "category": category,
        "index": index,
        "word" : word_to_add,
        "language":language
    }

    return new_index


def transform_dict_add_item(transforms_dicts,language,category,new_transform):
    # update db:
    dynamodb.update_transforms([new_transform])

    # updated cache :
    transforms_dict = dynamodb.get_transforms(language,category)
    transforms_dict_id = get_transforms_dict_id(language,category)
    transforms_dicts[transforms_dict_id] = transforms_dict


def create_transform_item(category,language,words_transforms,src_word,trg_word):
    id =  src_word + trg_word + language + category 
    hashed_id = hashlib.sha1(id.encode("UTF-8")).hexdigest()[5:]

    new_transform = {
        "id" : hashed_id,
        "category": category,
        "language": language,
        "sourceWord": src_word,
        "targetWord": trg_word
    }
    return new_transform

def create_product_word_item(user_info,index,product_word,start,file_duration,filename,end,category,assignmentId,language):
    
    _id = product_word + str(start) 
    #hashed_id = hashlib.sha1(_id.encode("UTF-8")).hexdigest()[5:]
    new_product_word = {
        #"id" : hashed_id,
        "index" : index,
        "word" : product_word,
        "start" : start,
        "language" : language,
        "category" : category,
        "totalDuration": file_duration
    }

    if(not (assignmentId == None) ) :
        new_product_word["assignmentId"] = assignmentId
        _id = _id +  assignmentId

    if(not (filename == None) ) :
        new_product_word["filename"] = filename
    
    if( (not (end == None)) and 
        (DynamoDB.version == "V2") ) :
        new_product_word["end"] = end

    if(not (user_info == None) ) :
        new_product_word["userInfo"] = user_info
        _id = _id + user_info
    
    _id += category
    _id += language

    hashed_id = hashlib.sha1(_id.encode("UTF-8")).hexdigest()[5:]
    new_product_word["id"] = hashed_id
    
    return new_product_word

def get_transforms_dict(transforms_dicts,language,category):
    transforms_dict_id = get_transforms_dict_id(language,category)
     
    if(transforms_dict_id in transforms_dicts):
        transforms_dict = transforms_dicts[transforms_dict_id]
    else:
        transforms_dict = dynamodb.get_transforms(language,category)
        transforms_dicts[transforms_dict_id] = transforms_dict
    
    return transforms_dict

def update_new_index(index_dicts,category,language,words_index,word):
    print('index does not contain :'+rev(word,language))
    print('creating new index ...')

    # create new index:
    new_index_item = create_index_item(category,language,words_index["nextIndexVal"],word)
    index_dict_add_last_item(index_dicts,language,category,new_index_item)
    
    index = new_index_item["id"]
    print('index:'+ str(index) +' created for word: '+word)
    return (index,word)

# TBD: remove redundunt words_transforms , words_index from functions calls
def update_new_transform(index_dicts,category,language,words_transforms,src_word,trg_word):
    print('creating new transform from :' + rev(src_word,language) + ' to:'+rev(trg_word ,language) ) 

    # create new transform:
    new_transform_item = create_transform_item(category,language,words_transforms,src_word,trg_word)
    transform_dict_add_item(transforms_dicts,language,category,new_transform_item)
    print('transform created.')

def updateAssignmentTranscriptDone(assignmentId):
    pass

def keyExists(dict, key):   
    if key in dict.keys(): 
        return True 
    else: 
        return False


# MAIN:
index_dicts = {}
transforms_dicts = {}
ask_for_translation = False

# variables:
#input_filename = "amaril-product-words-converter/input-normelizer/target.json"


# load input data:
#input_data = get_input(input_filename )
#input_data = validate_input(input_data)

# convert prrat:
 
def convert_prrat(filename,language,category):
    input_data = normelize_prrat(filename,language,category)
    input_json_str = json.dumps(input_data)
    input_obj = json.loads(input_json_str, parse_float=Decimal)
    is_prat = True
    push(input_obj,is_prat)

def convert_sonix_data(data,language,category):
    input_data = normelize_sonix_data(data,language,category)
    input_json_str = json.dumps(input_data)
    input_obj = json.loads(input_json_str, parse_float=Decimal)
    push(input_obj)

def process_praat_dir(input_dir, language,category):
    for current_file in os.listdir(input_dir):
        if current_file.endswith(".xlsx"):
            print('proocsseing file:',current_file)
            convert_prrat(input_dir+current_file,language,category)
    print('done.')
    return

def update_raw_word_product_items(raw_word,new_index,new_word,language,category):
    # get all word_items
    old_product_word_items = dynamodb.get_product_raw_word_items(raw_word,category)
    
    # update product items (with new words, and new indexes)
    new_product_word_items = []
    for old_product_word_item in old_product_word_items:
        new_product_word_item = old_product_word_item
        new_product_word_item["index"] = new_index
        new_product_word_item["word"] = new_word
        new_product_word_items.append(new_product_word_item)
    
    dynamodb.update_product_words(new_product_word_items)



def update_word_product_items(old_word,new_index,new_word,language,category):
    # get all word_items
    old_product_word_items = dynamodb.get_product_word_items(old_word,language,category)
    
    # update product items (with new words, and new indexes)
    new_product_word_items = []
    for old_product_word_item in old_product_word_items:
        new_product_word_item = old_product_word_item
        new_product_word_item["index"] = new_index
        new_product_word_item["word"] = new_word
        new_product_word_items.append(new_product_word_item)
    
    dynamodb.update_product_words(new_product_word_items)



#############################################################

# Converter 2.0:

def deconstIndexItem(item):
    transform = None
    index = None
    word = None
    if "transform" in item:
        transform = item["transform"]
    if "index" in item:
        index = item["index"]
    if "word" in item:
        word = item["word"]
    return (index,word,transform)

def getIndexItem(items,word):
    for item in items:
        if( item["word"] == word ):
            return item
    return None


def get_last_index_item(language,category):
    # get updated index:
    index_items = dynamodb.get_index_items(language,category)
    # function that filters index not existing
    def filterIndexes(item):
        return ('index' in item)

    filtered_index_item = list(filter(filterIndexes, index_items))

    # sort and filter items that are pure transforms :
    sorted_index_items = sorted(filtered_index_item, key=lambda k: k['index'])
    
    # get highest index :
    lastIndexItem = sorted_index_items[len(sorted_index_items)-1]
    
    return  lastIndexItem

def get_new_index(language,category):
    # get updated index:
    index_items = dynamodb.get_index_items(language,category)
    # function that filters index not existing
    def filterIndexes(item):
        return ('index' in item)

    filtered_index_items = list(filter(filterIndexes, index_items))
    return len(filtered_index_items)-1

"""
    we need to be able to get products according to raw_word !!
    dynamodb.get_product_word_items(word,category)    
    in order to update ONLY products that were transformed according to the speciifc transform we removed.
    (there could be other transforms which we dont want to add as the raw word with new index but keep as the transform)
"""
def removeTransform(item):

    # add new index to Item (with index = lastIndex+1)
    new_index = get_new_index(item["language"],item["category"])
    item["index"] = new_index
    
    # remove tranform from item:
    item.pop('transform', None)

    # update Index Item:
    dynamodb.update_index([item])

    # Apply to Products:
    #update_raw_word_product_items(item['word'],item["index"],item["word"],item["language"],item["category"])

  

def addTransform(item,transform):
    language = item['language']
    category = item['category']
    
    last_index_item = get_last_index_item(language,category)

    if(not last_index_item ):
        raise AssertionError('Error ,get_last_index_item failed.')

    # (item: index removed --> transform)
    # (last item : index --> index removed)
    index_removed = item.pop('index', None)
    if(not index_removed):
        # is item a transform ? (change existing transform)
        if('transform' in item):
            # change existing transform and exit:
            print('changing exsiting transform from: '+item['word']+'->'+item['transform']+' to: '+item['word']+'->'+transform)
            item['transform'] = transform
            dynamodb.update_index([item])
            return
        else:
            raise AssertionError('Warning: add transform: index nor transform props dont exist on item - doing nothing. item:'+str(item))
    
    item['transform'] = transform

    # are we adding transform to last item ?
    if(item['id'] == last_index_item['id']):
        dynamodb.update_index([item])
    else:
        last_index_item['index'] = index_removed
        item["transactionAction"] = 'Put'
        last_index_item["transactionAction"] = 'Put'
        dynamodb.atomic_update_index([item,last_index_item])


def verify_index(items,language,category):
    dynamodb.raw_index_to_dict(items,language,category)
    return

def handleRemoveTransform(word,language,category):

    # get updated index:
    index_items = dynamodb.get_index_items(language,category)

    # verify:
    verify_index(index_items,language,category)

    # get item:
    index_item = getIndexItem(index_items,word)
    if(not index_item):
        print('Error, item for '+word+' doesnt exist , do nothing and exit ...')
        return False
        
    # transform exists ? 
    if('transform'in index_item):
        removeTransform(index_item)
    else:
        print('transform for :'+word+' does not exist - no need to remove transform.')


def transforms_xlsx_to_ar(filename):

    wb1 = load_workbook(filename, data_only=True)
    item_ar = []
    sheet =  wb1.worksheets[0]
    for row_cells in sheet.iter_rows():
        key = row_cells[1].value        
        val = row_cells[0].value
        if((not key or not val or key == 'num') and not key==0 ):
            continue
        
        category = row_cells[1].value 
        language = row_cells[2].value 
        source_word = row_cells[3].value 
        target_word = row_cells[4].value

        item = {
            "category":category,
            "language":language,
            "sourceWord":source_word,
            "targetWord":target_word,
        }
        
        item_ar.append(item)
        
    return item_ar

def handleDeleteProductWords(filename,table_name):
    items = dynamodb.get_product_word_items_by_filename(filename)
    dynamodb.delete_items(table_name,items)
    return


def transform_product_words(pw_items,index_items):

    # sanity:
    if pw_items and len(pw_items) > 0:
        print('transform_product_words recieved'+str(len(pw_items))+'pw items')
    else:
        print('error, transform_product_words no pw items recieved.')
        return []
    
    if (not index_items ) or not len(index_items):
        print('error, transform_product_words no Index items recieved.')
        return []
    
    for pw_item in pw_items:
        # swap indexID for number:
        index_item = getIndexItem(index_items,pw_item["word"])
        if(not index_item):
            print('Error, item for '+pw_item["word"]+' doesnt exist index will be empty ...')
            pw_item.pop("index")
            continue
        else:
            # is transform ?
            if( (not ("index" in index_item) ) and ("transform" in index_item) ):
                transform_item = getIndexItem(index_items,index_item["transform"])
                if(not transform_item):
                    print('Error, item for '+index_item["transform"]+' doesnt exist index will be empty ...')
                    pw_item.pop("index")
                    continue
                pw_item["index"] = transform_item["index"]
                pw_item["word"] = transform_item["word"]
                pw_item["originalWord"] = index_item["word"]
            else:
                pw_item["index"] = index_item["index"]

    return pw_items

def handleGetProductWords(filename):
    pw_items = dynamodb.get_product_word_items_by_filename(filename)
    index_items = dynamodb.get_index_items(pw_items[0]["language"],pw_items[0]["category"])
    transformed_pw_items = transform_product_words(pw_items,index_items)
    return transformed_pw_items


def handleGetProductWordsByCategory(category,language):
    pw_items = dynamodb.get_product_word_items_by_category(category,language)
    index_items = dynamodb.get_index_items(pw_items[0]["language"],pw_items[0]["category"])
    transformed_pw_items = transform_product_words(pw_items,index_items)
    return transformed_pw_items

def handleAddTransforms(source_filename):
    transforms_ar = transforms_xlsx_to_ar(source_filename)
    for trans_item in transforms_ar:
        handleAddTransform(trans_item["sourceWord"],trans_item["targetWord"],trans_item["language"],trans_item["category"])

def handleAddTransform(word,transform,language,category):

    # get updated index:
    index_items = dynamodb.get_index_items(language,category)

    # verify index:
    verify_index(index_items,language,category)
   
    # normelize:
    word_normelized = normelize_word( word, language, category )
    transform_normelized = normelize_word( transform, language, category )

    # verify - source word is not other words tranform target meaning only 1 transform deep. 
    for index_item in index_items:
        if( ("transform" in index_item) and (index_item["transform"] == word_normelized) ):
            raise AssertionError('Error, handleAddTransform: we dont allow transformations of depth greater than 1 (already exists a tranfsorm: ' +index_item['word'] + '-->' + word_normelized )

    # get word item:
    index_item = getIndexItem(index_items,word_normelized)
    if(not index_item):
        print('Error, item for '+word_normelized+' doesnt exist , do nothing and exit ...')
        return False

    # get transform item:
    transform_item = getIndexItem(index_items,transform_normelized)
    if(not transform_item):
        print(transform_normelized + ',does not exist in index items ...')
        
        # create new index for transfom:
        index_dict = dynamodb.raw_index_to_dict(index_items,language,category)
        transform_item = create_index_item(category,language,index_dict["nextIndexVal"],transform_normelized)
        
        # update transform_item in DB:
        dynamodb.update_index([transform_item])

        print('created new index:'+rev(transform_normelized,language) + ' : ' + str(transform_item["index"]))
        
    if(not ("index" in transform_item) ):
        raise AssertionError('Error, handleAddTransform: we dont allow transformations of depth greater than 1 (no index prop for transform target item) - word:'+transform_normelized)

    addTransform(index_item,transform_normelized)


def index_verify_start(index_dict,language,category):
    # check start exists:
    if(not ("start" in  index_dict)):
        # create start index item :
        new_index_item = create_index_item(category,language,-1,'start')
        # override with -1:
        index_dict_add_last_item(index_dicts,language,category,new_index_item)
        print('index:'+ str(new_index_item["index"]) +' created for word: '+'start')
 


def push(input_data,is_prrat=False):

    recordings = input_data["recordings"] 
    for rec in recordings:
        # RecordingItem: TBD: create class  to validate etc ..
        (language,assignmentId,items) = (rec["language"],rec["assignmentId"],rec["items"])
        filename = None 
        if("filename" in rec.keys()):
            filename = rec["filename"]
        user_info = None
        if(is_prrat):
            user_info = assignmentId
            # create new a_id for praat files:
            a_id =  user_info + 'praat'
            hashed_id = hashlib.sha1(a_id.encode("UTF-8")).hexdigest()[5:]
            assignmentId = hashed_id

        for item in items:

            # WordItem:
            (file_duration,start,category,word,end ) = (round(item["file_duration"], 3), round(item["start"], 3),item["category"],item["word"],None ) 
            #print("file_duration",file_duration)
            if(keyExists(item, "end")):
                end = round(item["end"],3)
            # get dictionaries:
            words_index = get_index_dict(index_dicts,language,category)

            # verify start:
            index_verify_start(words_index,language,category)
            # we pull index again since start might have been added:
            words_index = get_index_dict(index_dicts,language,category)

            # NOTE: we normlize since index keys are all normelized
            word_normelized = normelize_word( word, language, category )
               
            # word in index ?
            if(word_normelized in words_index.keys()):
                # done:
                index = words_index[word_normelized]
            else:
                # create new index for word:
                (index,productWord) = update_new_index(index_dicts,category,language,words_index,word_normelized)
                print('created new index:'+rev(word_normelized,language) + ' : ' + str(index))

            # create ProductWordItem and write to db:
            new_product_word = create_product_word_item(user_info,index,word_normelized,start,file_duration,filename,end,category,assignmentId,language)
            dynamodb.update_product_words([new_product_word])
    
        # TBD: update assignment->transcriptDone to true 
        updateAssignmentTranscriptDone(assignmentId)
    return 

def handle_delete_from_index(index_word,language,category):
    # TBD: check that no transforms are leaning on this index ...
    print ('handle_delete_from_index:',index_word,category,language)
    
    # sanity (do all checks here before we modify productwords):
    index_item = dynamodb.get_index_item_by_word(index_word,language,category)
    
    if(not index_item):
        raise AssertionError('Error: handle_delete_from_index: '+ index_word +' does not exist in index.')

    if('transform' in index_item):
        raise AssertionError('Warning: trying to delete a transform .. (please first removeTransform and then try to delete)'+str(index_item))

    if( not ('index' in index_item) ):
        raise AssertionError('Warning: trying to delete an item that has no index column '+str(index_item))

    last_index_item = get_last_index_item(index_item['language'],index_item['category'])
    if(not last_index_item ):
        raise AssertionError('Error ,get_last_index_item failed.')


    # (1) process productWords:
    product_word_items = dynamodb.get_product_word_items(index_word,language,category)
    for p_word_item in product_word_items:
        if(not ('assignmentId' in p_word_item)):
            raise AssertionError('Error ,handleDeleteFromIndex failed. no assignmentId on index item')

        assignment_id = p_word_item['assignmentId']
        
        # delete pw item:
        dynamodb.delete_items(dynamodb.get_table("productWords"),[p_word_item])

    print ('handle_delete_from_index: Removed all occurrences of: '+ index_word +' in ProductWords table ... removing now from Index Table ..')

    # (2) Delete - the index_item from the index and close the hole created (last item : index --> index removed):
    index_removed = index_item.pop('index', None)

    # are we deleting last item ?
    if(index_item['id'] == last_index_item['id']):
        dynamodb.delete_items(dynamodb.get_table("index"),[index_item])
    else:
        last_index_item['index'] = index_removed
        last_index_item["transactionAction"] = 'Put'
        index_item["transactionAction"] = 'Delete'
        dynamodb.atomic_update_index([index_item,last_index_item])

    print ('handle_delete_from_index: Success for word '+ index_word )



def handle_get_index(language,category,target_filename):
    index_items = dynamodb.get_index_items(language,category)
    update_index_sheet(target_filename,index_items)